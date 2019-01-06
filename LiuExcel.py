import openpyxl
from openpyxl import Workbook


def load_workbook(name:str)->openpyxl.Workbook:
    '''
    :param name: string include path file
    :return: workbook from openpxl
    '''

    return openpyxl.load_workbook(name)



# my_workbook = load_workbook("mergeCell.xlsx")
#
# list_of_sheet_name = my_workbook.sheetnames
#
# sheet = my_workbook[list_of_sheet_name[0]]

def search_case_insensitive(string_for_searching:str, workbook:Workbook,sheet_name:str, column:list)->list:
    '''
    find rows that have match string (not substhing)

    :param string_for_searching: specific string for search
    :param workbook: workbook from openpxl
    :param sheet_name: sheet name from work book
    :param column: specific column in work book
    :return: list of row
    '''
    sheet = workbook[sheet_name]
    my_list = []

    my_set = set()
    for column_name in column:

        #get a column in excel
        local_cell_tuple = sheet[column_name]

        for cell in local_cell_tuple:

            # compare
            if string_for_searching.lower() == str(cell.value).lower():

                # get specific row from sheet
                row = sheet[cell.row]

                # check duplicate row
                result: set = my_set & {cell.row}
                if result.__len__() == 0:
                    my_list.append(row)

                    # update set for check duplicate next time
                    my_set = result

    return my_list


def search_case_sensitive(string_for_searching: str, workbook: Workbook, sheet_name: str, column: list) -> list:
    '''
    find rows that have match string (not substhing)

    :param string_for_searching: specific string for search
    :param workbook: workbook from openpxl
    :param sheet_name: sheet name from work book
    :param column: specific column in work book
    :return: list of row
    '''
    sheet = workbook[sheet_name]
    my_list = []

    my_set = set()
    for column_name in column:

        # get a column in excel
        local_cell_tuple = sheet[column_name]

        for cell in local_cell_tuple:

            # compare
            if string_for_searching == str(cell.value):

                # get specific row from sheet
                row = sheet[cell.row]

                # check duplicate row
                result: set = my_set & {cell.row}
                if result.__len__() == 0:
                    my_list.append(row)

                    # update set for check duplicate next time
                    my_set = result

    return my_list




def search_case_insensitive_all_sheet(string_for_searching: str, workbook: Workbook, sheet_name: str) -> list:

    '''
    seach case insensitive all sheet
    :param string_for_searching: string_for_searching: specific string for search
    :param workbook: workbook from openpxl
    :param sheet_name: sheet name from work book
    :return: list of dictionary
    '''
    sheet = workbook[sheet_name]

    max_row = sheet.max_row
    max_column = sheet.max_column

    my_list_for_return = []
    my_set = set()
    my_dictionary = {}

    # iter each cell in excel
    for row in range(1, max_row+1):
        for column in range(1, max_column):

            if string_for_searching.lower() == str(sheet.cell(row=row, column= column).value).lower():

                # get specific row from sheet
                row_in_sheet = sheet[row]

                # check duplicate row
                result: set = my_set & {row}
                if result.__len__() == 0:

                    coordinate = sheet.cell(row= row, column= column).coordinate
                    my_dictionary[coordinate] = row_in_sheet
                    my_list_for_return.append(my_dictionary)
                    my_dictionary = {}

                    # update set for check duplicate next time
                    my_set = result

    return my_list_for_return


def search_substring(string_for_searching:str, workbook:Workbook,sheet_name:str, column:list)->list:
    '''
    find rows that have match string (not substhing)

    :param string_for_searching: specific string for search
    :param workbook: workbook from openpxl
    :param sheet_name: sheet name from work book
    :param column: specific column in work book
    :return: list of row
    '''
    sheet = workbook[sheet_name]
    my_list = []

    my_set = set()
    for column_name in column:

        #get a column in excel
        local_cell_tuple = sheet[column_name]

        for cell in local_cell_tuple:

            # compare
            if string_for_searching in str(cell.value):

                # get specific row from sheet
                row = sheet[cell.row]

                # check duplicate row
                result: set = my_set & {cell.row}
                if result.__len__() == 0:
                    my_list.append(row)

                    # update set for check duplicate next time
                    my_set = result

    return my_list


